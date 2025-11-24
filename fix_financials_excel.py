from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from pathlib import Path

WB_PATH = Path("data/raport_finansowy_2024.xlsx")


def format_zestawienie(ws):
    headers = [cell.value for cell in ws[1]]
    col_map = {name: idx + 1 for idx, name in enumerate(headers)}
    required = ["koszty_operacyjne", "liczba_uczniow", "koszt_na_ucznia"]
    for r in required:
        if r not in col_map:
            raise ValueError(f"Brakuje kolumny {r} w Zbiorcze_porownanie")

    col_cost = col_map["koszty_operacyjne"]
    col_count = col_map["liczba_uczniow"]
    col_kpu = col_map["koszt_na_ucznia"]

    money_cols = [
        "przychody_netto",
        "dotacje_podstawowe",
        "przychody_budzetowe",
        "koszty_operacyjne",
        "amortyzacja",
        "materialy_i_energia",
        "uslugi_obce",
        "podatki_i_oplaty",
        "wynagrodzenia",
        "ubezpieczenia_i_swiadczenia",
        "pozostale_koszty_rodzajowe",
        "pozostale_przychody_operacyjne",
        "pozostale_koszty_operacyjne",
        "zysk_strata_netto",
        "koszt_na_ucznia",
    ]
    money_format = "#,##0.00"
    count_format = "0"

    for row in range(2, ws.max_row + 1):
        cost_addr = f"{get_column_letter(col_cost)}{row}"
        count_addr = f"{get_column_letter(col_count)}{row}"
        ws.cell(row=row, column=col_kpu).value = f'=IFERROR({cost_addr}/{count_addr}, "")'

    for name in money_cols:
        if name in col_map:
            col = col_map[name]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = money_format
    for name in ["liczba_uczniow"]:
        if name in col_map:
            col = col_map[name]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = count_format

    last_col = get_column_letter(ws.max_column)
    last_row = ws.max_row
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def rebuild_pivot_placowka(wb):
    src = wb["Zbiorcze_porownanie"]
    headers = [cell.value for cell in src[1]]
    col_map = {name: idx + 1 for idx, name in enumerate(headers)}
    needed = ["placowka", "koszty_operacyjne", "zysk_strata_netto", "liczba_uczniow"]
    for n in needed:
        if n not in col_map:
            raise ValueError(f"Brakuje kolumny {n} w Zbiorcze_porownanie")

    rows = []
    for r in range(2, src.max_row + 1):
        row = {h: src.cell(r, col_map[h]).value for h in needed}
        rows.append(row)
    rows.sort(key=lambda x: (x["koszty_operacyjne"] is None, -(x["koszty_operacyjne"] or 0)))

    if "Pivot_placowka" in wb.sheetnames:
        del wb["Pivot_placowka"]
    ws = wb.create_sheet("Pivot_placowka")
    headers_out = ["placowka", "koszty_operacyjne", "zysk_strata_netto", "liczba_uczniow", "koszt_na_ucznia"]
    ws.append(headers_out)
    for i, row in enumerate(rows, start=2):
        ws.cell(i, 1, row["placowka"])
        ws.cell(i, 2, row["koszty_operacyjne"])
        ws.cell(i, 3, row["zysk_strata_netto"])
        ws.cell(i, 4, row["liczba_uczniow"])
        ws.cell(i, 5, f'=IFERROR({get_column_letter(2)}{i}/{get_column_letter(4)}{i}, "")')

    money_format = "#,##0.00"
    count_format = "0"
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 2).number_format = money_format
        ws.cell(r, 3).number_format = money_format
        ws.cell(r, 5).number_format = money_format
        ws.cell(r, 4).number_format = count_format
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    last_col = get_column_letter(ws.max_column)
    last_row = ws.max_row
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    return ws


def rebuild_charts(wb, pivot_ws):
    if "Wykresy" in wb.sheetnames:
        del wb["Wykresy"]
    ws_chart = wb.create_sheet("Wykresy")
    max_row = pivot_ws.max_row
    categories = Reference(pivot_ws, min_col=1, min_row=2, max_row=max_row)

    def add_chart(col_idx, title, pos, y_title):
        data = Reference(pivot_ws, min_col=col_idx, min_row=1, max_row=max_row)
        chart = BarChart()
        chart.title = title
        chart.y_axis.title = y_title
        chart.x_axis.title = "Placówka"
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.dataLabels = DataLabelList(showVal=True)
        ws_chart.add_chart(chart, pos)

    add_chart(2, "Koszty operacyjne per placówka", "B2", "PLN")
    add_chart(3, "Wynik netto per placówka", "B22", "PLN")
    add_chart(5, "Koszt na ucznia per placówka", "B42", "PLN/uczeń")


def main():
    wb = load_workbook(WB_PATH)
    if "Zbiorcze_porownanie" not in wb.sheetnames:
        raise SystemExit("Brak arkusza Zbiorcze_porownanie w pliku.")
    format_zestawienie(wb["Zbiorcze_porownanie"])
    pivot_ws = rebuild_pivot_placowka(wb)
    rebuild_charts(wb, pivot_ws)
    wb.save(WB_PATH)
    print("Zaktualizowano raport_finansowy_2024.xlsx: koszt_na_ucznia, pivot per placówka, wykresy.")


if __name__ == "__main__":
    main()
